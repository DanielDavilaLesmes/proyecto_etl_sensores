# src/extract.py
import os
import re
from openpyxl import load_workbook
from src.config import obtener_configuracion_por_nombre_interno, obtener_celda_pasillo

def encontrar_archivos_por_procesar(input_folder):
    """
    Busca archivos .xlsx en la carpeta de entrada especificada.
    
    Args:
        input_folder (str): Ruta completa a la carpeta de entrada de los archivos Excel.

    Returns:
        list of str: Lista de rutas completas a los archivos Excel.
    """
    archivos_a_procesar = []
    # Usar la carpeta de entrada proporcionada por run_etl.py
    if not os.path.exists(input_folder):
        print(f"ERROR: La carpeta de entrada no existe: {input_folder}")
        return archivos_a_procesar

    # Todos los archivos .xlsx se consideran candidatos
    archivos_en_input = [f for f in os.listdir(input_folder) if f.endswith('.xlsx')]
    print(f"-> Encontrados {len(archivos_en_input)} archivos en la carpeta de entrada. Se determinará la configuración por contenido...")

    for filename in archivos_en_input:
        filepath = os.path.join(input_folder, filename)
        # Se agrega la ruta completa, la configuración se obtiene en la siguiente función
        archivos_a_procesar.append(filepath)
            
    return archivos_a_procesar

def leer_archivo_excel(filepath):
    """
    Lee datos de un archivo Excel.
    1. Lee la celda de metadatos (ej: 'B1') para identificar el Pasillo.
    2. Usa el Pasillo para encontrar la configuración correcta.
    3. Extrae cabeceras y datos según la configuración (data_start_row).
    
    Args:
        filepath (str): Ruta completa al archivo Excel.
        
    Returns:
        tuple (list, list of list, dict) o (None, None, None)
        (headers, data_rows, config_con_pasillo_resuelto)
    """
    filename = os.path.basename(filepath)
    print(f"\n--- EXTRACCIÓN: {filename} ---")
    
    try:
        # Cargar el workbook (solo lectura y sin datos)
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        sheet = workbook.active
        
        # 1. Identificar la celda donde está el nombre del Pasillo (ej: 'B1')
        celda_pasillo_key = obtener_celda_pasillo(filename)
        
        # Leer el valor del Pasillo
        nombre_interno_pasillo = sheet[celda_pasillo_key].value
        
        if not nombre_interno_pasillo:
            print(f"ERROR: No se encontró el nombre del Pasillo en la celda {celda_pasillo_key}. Saltando archivo.")
            return None, None, None

        # 2. Obtener la configuración basada en el nombre interno
        config = obtener_configuracion_por_nombre_interno(str(nombre_interno_pasillo))
        
        if not config:
            print(f"ERROR: No se encontró configuración para el Pasillo: '{nombre_interno_pasillo}'. Saltando archivo.")
            return None, None, None

        pasillo = config['nombre_pasillo']
        data_start_row = config.get('data_start_row', 1) 
        print(f"-> Archivo identificado como: {pasillo}. Cabeceras inician en Fila {data_start_row}.")

        
        # 3. EXTRAER CABECERAS Y DATOS
        
        headers = []
        data_rows = []
        
        # Identificar Cabeceras (en la fila definida por data_start_row)
        header_row_number = data_start_row
        
        for idx, row in enumerate(sheet.iter_rows(min_row=header_row_number, max_row=header_row_number, values_only=True), start=header_row_number):
            headers = [str(cell).strip() if cell is not None else f"Col_{i}" 
                       for i, cell in enumerate(row)]
            break # Solo necesitamos la primera fila (cabecera)

        # Extraer Filas de Datos (inician en la fila siguiente)
        start_data_row = header_row_number + 1
        
        for row in sheet.iter_rows(min_row=start_data_row, values_only=True):
            data_rows.append(list(row)) 
            
        print(f"-> Extracción exitosa. Filas leídas: {len(data_rows)} | Cabeceras encontradas: {len(headers)}")
        
        return headers, data_rows, config
        
    except FileNotFoundError:
        print(f"ERROR: Archivo no encontrado en la ruta: {filepath}")
        return None, None, None
    except Exception as e:
        print(f"ERROR al leer {filename}: {e}")
        return None, None, None